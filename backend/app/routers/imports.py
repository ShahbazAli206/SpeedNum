"""CSV / XLSX bulk import with a review step before anything is written.

Two importers share the file-reading and column-detection machinery below:
clients (the client book) and users (staff + client-portal logins). Both follow
the same two-call shape — POST .../preview returns a row-by-row validation
report and never writes, POST .../commit applies the rows the operator kept —
because an accounting firm's spreadsheet is always messier than it looks and a
half-applied import is worse than a rejected one.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from sqlalchemy import select

from ..deps import AdminUserDep, SessionDep, SuperadminDep, TenantUserDep, client_ip
from ..models import Client, Profile, Service, Tenant
from ..seats import remaining_client_seats, remaining_staff_seats
from ..schemas import (
    ImportCommitRequest,
    ImportPreview,
    ImportPreviewRow,
    ImportResult,
    TenantAdminCreate,
    TenantImportOutcome,
    TenantImportResult,
    UserImportResult,
    UserImportRow,
)
from ..services import accounts, audit
from ..services.accounts import AccountError
from ..services.rate_limit import rate_limit_by_tenant
from .admin import provision_tenant
from .clients import MIN_ANNUAL_FEE

router = APIRouter(prefix="/import", tags=["import"])

MAX_PREVIEW_ROWS = 100

# Bulk-provisioning logins is slow (one Supabase round-trip each) and mistakes
# are expensive to undo, so a single commit is capped.
MAX_USER_COMMIT_ROWS = 200

# Lower than team.py's/users.py's per-account limit — a single bulk commit
# can itself create up to MAX_USER_COMMIT_ROWS logins, so the *number of
# commit calls* needs a tighter cap than the number of individual creations.
_bulk_user_import_rate_limit = rate_limit_by_tenant("import-users-commit", limit=5, window_seconds=3600)

# Each row provisions an entire firm (tenant + admin login + welcome email),
# so the cap is far tighter than clients/users. No rate_limit_by_tenant here:
# it requires TenantUserDep, which 409s a superadmin with no tenant of their
# own — the SuperadminDep gate below is already the narrowest role on the
# platform, so a hard row cap is the only guard needed.
MAX_TENANT_COMMIT_ROWS = 25

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "code": ("client code", "code", "client id", "client no", "account", "account number"),
    "legal_name": ("legal name", "client name", "name", "company", "company name", "legal"),
    "business_name": ("operating name", "trade name", "business name", "dba", "business"),
    "client_type": ("type", "entity type", "client type"),
    "status": ("status", "client status"),
    # "primary contact email/phone" is what the downloadable template calls
    # these, so the template's own headers have to round-trip.
    "email": (
        "email", "e-mail", "email address", "contact email",
        "primary contact email", "primary email", "main email",
    ),
    "phone": (
        "phone", "telephone", "phone number", "contact phone", "mobile",
        "primary contact phone", "primary phone", "main phone",
    ),
    "business_number": ("bn", "business number", "cra bn", "cra business number", "sin"),
    "gst_number": ("gst", "gst number", "gst/hst", "hst number", "gst/hst number"),
    "payroll_number": ("payroll", "payroll number", "rp account"),
    "address_line1": ("address", "address 1", "street", "address line 1", "mailing address"),
    "address_line2": ("address 2", "address line 2", "suite", "unit"),
    "city": ("city", "town"),
    "province": ("province", "state", "prov"),
    "postal_code": ("postal code", "postal", "zip", "zip code"),
    "fiscal_year_end": ("fiscal year end", "year end", "fye", "year-end", "fiscal year-end"),
    "year_end_month": ("year end month", "fye month"),
    "year_end_day": ("year end day", "fye day"),
    "annual_fee": ("annual fee", "fee", "fees", "billing", "annual billing"),
    # Firms often export the monthly figure rather than the annual one the
    # database stores — converted to annual_fee (×12) in parse_row.
    "mrr": ("mrr", "monthly recurring revenue", "monthly fee", "monthly billing", "recurring revenue"),
    # Not a real column (see new-client-client.tsx) — the plan lives in
    # `tags`, so this just gets folded in there.
    "plan": ("plan", "package", "tier"),
    # Resolved to owner_id by matching against the tenant's team roster
    # (parse_row / _team_lookup below); a name with no match just leaves the
    # client unassigned rather than rejecting the row.
    "owner": ("owner", "accountant", "manager", "assigned to", "accountant / manager", "assigned accountant"),
    "notes": ("notes", "comments", "remarks"),
    "tags": ("tags", "labels", "groups"),
}

CLIENT_TYPE_MAP = {
    "corporation": "corporation", "corp": "corporation", "inc": "corporation",
    "incorporated": "corporation", "ltd": "corporation", "company": "corporation",
    "sole proprietor": "sole_proprietor", "sole proprietorship": "sole_proprietor",
    "sole prop": "sole_proprietor", "proprietorship": "sole_proprietor",
    "partnership": "partnership", "llp": "partnership",
    "individual": "individual", "personal": "individual", "t1": "individual", "person": "individual",
    "nonprofit": "nonprofit", "non-profit": "nonprofit", "npo": "nonprofit", "charity": "nonprofit",
    "trust": "trust", "estate": "trust",
}

STATUS_MAP = {
    "active": "active", "current": "active", "client": "active",
    "prospect": "prospect", "lead": "prospect", "pending": "prospect",
    "inactive": "inactive", "dormant": "inactive", "former": "inactive",
    "archived": "archived", "closed": "archived",
}

MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _normalise(header: str) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", header.strip().lower()).strip()


def detect_mapping(columns: list[str]) -> dict[str, str]:
    """Map source column -> client field, best guess."""
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    for column in columns:
        key = _normalise(column)
        for field, aliases in FIELD_ALIASES.items():
            if field in taken:
                continue
            if key == field or key in aliases:
                mapping[column] = field
                taken.add(field)
                break
    return mapping


def _parse_year_end(raw: str) -> tuple[int, int] | None:
    value = raw.strip().lower()
    if not value:
        return None

    iso = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", value)
    if iso:
        return int(iso.group(2)), int(iso.group(3))

    pair = re.match(r"^(\d{1,2})[-/](\d{1,2})$", value)
    if pair:
        first, second = int(pair.group(1)), int(pair.group(2))
        return (first, second) if first <= 12 else (second, first)

    named = re.match(r"^([a-z]{3,9})[ .\-]*(\d{1,2})$", value)
    if named:
        month = MONTH_NAMES.get(named.group(1)[:3])
        if month:
            return month, int(named.group(2))

    named_rev = re.match(r"^(\d{1,2})[ .\-]*([a-z]{3,9})$", value)
    if named_rev:
        month = MONTH_NAMES.get(named_rev.group(2)[:3])
        if month:
            return month, int(named_rev.group(1))
    return None


def _to_float(raw: str) -> float | None:
    cleaned = re.sub(r"[^0-9.\-]", "", raw or "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_row(
    raw: dict[str, Any],
    mapping: dict[str, str],
    owners_by_name: dict[str, str] | None = None,
) -> tuple[dict[str, Any], list[str]]:
    data: dict[str, Any] = {}
    errors: list[str] = []
    # `tags` and `plan` both write here rather than straight into `data`, since
    # either can appear alone or together and one must not clobber the other
    # depending on which column happens to come first in the file.
    tags: list[str] = []
    # A column the alias table doesn't recognise at all isn't dropped — it
    # lands here, the same free-form field the Add Client page's "Additional
    # information" section writes to, so a firm's own spreadsheet quirks (a
    # provincial corporation number, an internal code, an accountant name that
    # matched nobody, ...) survive the import instead of being silently
    # discarded or blocking the row outright.
    custom: dict[str, str] = {}

    for column, field in mapping.items():
        value = raw.get(column)
        text = "" if value is None else str(value).strip()
        if not text:
            continue

        if field == "fiscal_year_end":
            parsed = _parse_year_end(text)
            if parsed is None:
                errors.append(f"Could not read a year end from '{text}'")
            else:
                data["year_end_month"], data["year_end_day"] = parsed
        elif field in ("year_end_month", "year_end_day"):
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a number for {field.replace('_', ' ')}")
            else:
                data[field] = int(number)
        elif field == "annual_fee":
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a valid fee")
            else:
                data[field] = number
        elif field == "mrr":
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a valid MRR")
            else:
                # The database only stores the annual figure (see
                # new-client-client.tsx's own "Annual fee ($)" field) — the
                # client list then derives the monthly figure back out of it.
                data["annual_fee"] = round(number * 12, 2)
        elif field == "plan":
            tags.append(text)
        elif field == "owner":
            resolved = (owners_by_name or {}).get(text.lower())
            if resolved is None:
                # Not a blocking error, unlike an unmatched client name on the
                # user importer: that changes which *kind* of account gets
                # created (portal vs. staff), where this just leaves a client
                # unassigned — routine before a firm's roster is fully
                # populated, and easy to fix by hand afterwards. The name is
                # kept, not discarded, so nothing has to be re-typed later.
                custom[column] = text
            else:
                data["owner_id"] = resolved
        elif field == "client_type":
            data[field] = CLIENT_TYPE_MAP.get(text.lower(), "corporation")
        elif field == "status":
            data[field] = STATUS_MAP.get(text.lower(), "active")
        elif field == "tags":
            tags.extend(tag.strip() for tag in re.split(r"[,;|]", text) if tag.strip())
        elif field == "email":
            if "@" not in text:
                errors.append(f"'{text}' is not a valid email")
            else:
                data[field] = text.lower()
        else:
            data[field] = text

    if tags:
        data["tags"] = tags

    for column, value in raw.items():
        if column in mapping:
            continue
        text = "" if value is None else str(value).strip()
        if text:
            custom[column] = text
    if custom:
        data["custom"] = custom

    if not data.get("legal_name"):
        errors.append("Legal name is required")

    month = data.get("year_end_month", 12)
    day = data.get("year_end_day", 31)
    if not 1 <= int(month) <= 12:
        errors.append("Year end month must be between 1 and 12")
        data["year_end_month"] = 12
    else:
        try:
            date(2024, int(month), int(day))
        except ValueError:
            errors.append(f"{month}/{day} is not a real date")
            data["year_end_day"] = 28

    return data, errors


async def _read_table(upload: UploadFile) -> tuple[list[str], list[dict[str, Any]]]:
    raw = await upload.read()
    if not raw:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The uploaded file is empty.")
    filename = (upload.filename or "").lower()

    if filename.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is declared
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "Excel support is unavailable on this server. Please upload a CSV instead.",
            ) from exc

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The spreadsheet has no rows.") from exc
        columns = [str(cell).strip() if cell is not None else f"column_{i}" for i, cell in enumerate(header)]
        records = [
            {columns[i]: cell for i, cell in enumerate(row) if i < len(columns)}
            for row in rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        workbook.close()
        return columns, records

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 always decodes
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Could not decode the file.")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    columns = [c for c in (reader.fieldnames or []) if c]
    if not columns:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "No header row was found.")
    records = [row for row in reader if any((v or "").strip() for v in row.values() if isinstance(v, str))]
    return columns, records


@router.post("/clients/preview", response_model=ImportPreview)
async def preview_clients(
    session: SessionDep, user: TenantUserDep, file: UploadFile = File(...)
) -> ImportPreview:
    columns, records = await _read_table(file)
    mapping = detect_mapping(columns)
    if "legal_name" not in mapping.values():
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "No client-name column was recognised. Rename a column to 'Legal Name' and try again.",
        )

    owners_by_name = await _team_lookup(session, user.tenant_id)

    rows: list[ImportPreviewRow] = []
    valid = 0
    for index, record in enumerate(records[:MAX_PREVIEW_ROWS], start=2):
        data, errors = parse_row(record, mapping, owners_by_name)
        if not errors:
            valid += 1
        rows.append(ImportPreviewRow(row=index, data=data, errors=errors))

    return ImportPreview(
        columns=columns,
        detected_mapping=mapping,
        rows=rows,
        total_rows=len(records),
        valid_rows=valid,
    )


@router.post("/clients/commit", response_model=ImportResult)
async def commit_clients(
    payload: ImportCommitRequest, session: SessionDep, user: TenantUserDep
) -> ImportResult:
    created = updated = failed = 0
    errors: list[str] = []
    allowed = {column.name for column in Client.__table__.columns}

    # Tracked locally rather than re-querying COUNT(*) before every row — see
    # seats.remaining_client_seats. Only decremented for genuinely new rows
    # (existing.None branch below); an update never touches the cap.
    seats_left = await remaining_client_seats(session, user.tenant)

    for index, row in enumerate(payload.rows, start=1):
        data = {k: v for k, v in row.items() if k in allowed and k not in ("id", "tenant_id")}
        legal_name = (data.get("legal_name") or "").strip()
        if not legal_name:
            failed += 1
            errors.append(f"Row {index}: missing legal name")
            continue

        fee = data.get("annual_fee")
        if fee is not None and 0 < fee < MIN_ANNUAL_FEE:
            failed += 1
            errors.append(f"Row {index} ({legal_name}): annual fee must be at least ${MIN_ANNUAL_FEE}, or 0")
            continue

        existing = None
        if payload.update_existing:
            if data.get("code"):
                existing = await session.scalar(
                    select(Client).where(
                        Client.tenant_id == user.tenant_id, Client.code == data["code"]
                    )
                )
            if existing is None:
                existing = await session.scalar(
                    select(Client).where(
                        Client.tenant_id == user.tenant_id, Client.legal_name == legal_name
                    )
                )

        if existing is None and seats_left is not None and seats_left <= 0:
            failed += 1
            errors.append(f"Row {index} ({legal_name}): client seat limit reached — contact your provider for more seats")
            continue

        try:
            # A savepoint keeps one bad row from discarding the whole import.
            async with session.begin_nested():
                if existing is not None:
                    for key, value in data.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    session.add(Client(tenant_id=user.tenant_id, created_by=user.profile.id, **data))
                    created += 1
                    if seats_left is not None:
                        seats_left -= 1
        except Exception as exc:  # noqa: BLE001 - surface the row that failed
            if existing is not None:
                updated -= 1
            else:
                created -= 1
                if seats_left is not None:
                    seats_left += 1
            failed += 1
            errors.append(f"Row {index} ({legal_name}): {type(exc).__name__}")

    await audit.record(
        session,
        tenant_id=user.tenant_id,
        actor_id=user.profile.id,
        actor_email=user.profile.email,
        action="imported",
        entity="client",
        summary=f"Imported {created} new and updated {updated} client(s)",
        metadata={"created": created, "updated": updated, "failed": failed},
    )
    return ImportResult(created=created, updated=updated, failed=failed, errors=errors[:25])


# -----------------------------------------------------------------------------
# Users — staff and client-portal logins from a spreadsheet
# -----------------------------------------------------------------------------
USER_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "full_name": ("full name", "name", "user", "person", "employee", "accountant", "staff name"),
    "email": ("email", "e-mail", "email address", "work email", "login", "username"),
    "role": ("role", "access", "permission", "access level", "user role"),
    "title": ("title", "job title", "position", "designation"),
    "phone": ("phone", "telephone", "phone number", "mobile", "cell"),
    "client": ("client", "client name", "linked client", "company", "portal client"),
}

# What a firm is likely to write in a "role" column, mapped onto our four roles.
ROLE_MAP = {
    "owner": "owner", "partner": "owner", "principal": "owner",
    "admin": "admin", "administrator": "admin", "manager": "admin", "office manager": "admin",
    "member": "member", "staff": "member", "accountant": "member", "cpa": "member",
    "bookkeeper": "member", "associate": "member", "user": "member",
    "viewer": "viewer", "read only": "viewer", "read-only": "viewer", "guest": "viewer",
    # A client-portal login is expressed by naming a client, not by this column;
    # accepting the word here keeps the sheet readable and is harmless.
    "client": "member", "portal": "member",
}


def detect_user_mapping(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    for column in columns:
        key = _normalise(column)
        for field, aliases in USER_FIELD_ALIASES.items():
            if field in taken:
                continue
            if key == field or key == field.replace("_", " ") or key in aliases:
                mapping[column] = field
                taken.add(field)
                break
    return mapping


def parse_user_row(
    raw: dict[str, Any], mapping: dict[str, str], clients_by_name: dict[str, str]
) -> tuple[dict[str, Any], list[str]]:
    data: dict[str, Any] = {"role": "member"}
    errors: list[str] = []

    for column, field in mapping.items():
        value = raw.get(column)
        text = "" if value is None else str(value).strip()
        if not text:
            continue

        if field == "email":
            lowered = text.lower()
            if "@" not in lowered or "." not in lowered.split("@")[-1]:
                errors.append(f"'{text}' is not a valid email")
            else:
                data["email"] = lowered
        elif field == "role":
            resolved = ROLE_MAP.get(text.lower())
            if resolved is None:
                errors.append(f"'{text}' is not a role we recognise — defaulting to member")
                data["role"] = "member"
            else:
                data["role"] = resolved
        elif field == "client":
            matched = clients_by_name.get(text.strip().lower())
            if matched is None:
                errors.append(f"No client matches '{text}' — the account will be firm staff")
            else:
                data["client_id"] = matched
        else:
            data[field] = text

    if not data.get("email"):
        errors.append("Email is required")
    if not data.get("full_name"):
        # Derive something usable rather than rejecting the row outright: the
        # local part of the address is a better default than a blank name.
        local = str(data.get("email", "")).split("@")[0]
        derived = local.replace(".", " ").replace("_", " ").replace("-", " ").strip().title()
        if derived:
            data["full_name"] = derived
        else:
            errors.append("Full name is required")

    return data, errors


async def _client_lookup(session: SessionDep, tenant_id: Any) -> dict[str, str]:
    """Every way a spreadsheet might name one of this firm's clients."""
    rows = (
        await session.execute(
            select(Client.id, Client.legal_name, Client.business_name, Client.code).where(
                Client.tenant_id == tenant_id
            )
        )
    ).all()
    lookup: dict[str, str] = {}
    for client_id, legal_name, business_name, code in rows:
        for label in (legal_name, business_name, code):
            if label:
                lookup[str(label).strip().lower()] = str(client_id)
    return lookup


async def _team_lookup(session: SessionDep, tenant_id: Any) -> dict[str, str]:
    """Every firm-staff full name, for resolving a client import's
    accountant/owner column to `owner_id`."""
    rows = (
        await session.execute(
            select(Profile.id, Profile.full_name).where(
                Profile.tenant_id == tenant_id, Profile.client_id.is_(None)
            )
        )
    ).all()
    return {str(full_name).strip().lower(): str(profile_id) for profile_id, full_name in rows if full_name}


@router.post("/users/preview", response_model=ImportPreview)
async def preview_users(
    session: SessionDep, user: AdminUserDep, file: UploadFile = File(...)
) -> ImportPreview:
    """Validate a user spreadsheet without creating a single login."""
    columns, records = await _read_table(file)
    mapping = detect_user_mapping(columns)
    if "email" not in mapping.values():
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "No email column was recognised. Rename a column to 'Email' and try again.",
        )

    clients_by_name = await _client_lookup(session, user.tenant_id)
    existing_emails = {
        row.lower()
        for row in (
            await session.scalars(select(Profile.email).where(Profile.tenant_id == user.tenant_id))
        ).all()
    }

    rows: list[ImportPreviewRow] = []
    valid = 0
    seen: set[str] = set()
    for index, record in enumerate(records[:MAX_PREVIEW_ROWS], start=2):
        data, errors = parse_user_row(record, mapping, clients_by_name)
        email = str(data.get("email", ""))
        if email:
            if email in existing_emails:
                errors.append("An account already exists for this email")
            elif email in seen:
                errors.append("This email is duplicated earlier in the file")
            seen.add(email)
        if not errors:
            valid += 1
        rows.append(ImportPreviewRow(row=index, data=data, errors=errors))

    return ImportPreview(
        columns=columns,
        detected_mapping=mapping,
        rows=rows,
        total_rows=len(records),
        valid_rows=valid,
    )


@router.post(
    "/users/commit",
    response_model=UserImportResult,
    dependencies=[Depends(_bulk_user_import_rate_limit)],
)
async def commit_users(
    rows: list[UserImportRow],
    session: SessionDep,
    user: AdminUserDep,
    send_email: bool = True,
) -> UserImportResult:
    """Provision a login per row and email each person their credentials.

    Every row is attempted independently: one duplicate email or one Supabase
    rejection reports itself and the rest still land, because re-running a
    partially applied import would trip over the accounts that did succeed.
    """
    if len(rows) > MAX_USER_COMMIT_ROWS:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Import at most {MAX_USER_COMMIT_ROWS} users at a time — split the file.",
        )

    outcomes: list[Any] = []
    created = failed = emailed = 0
    errors: list[str] = []

    # Only staff rows (client_id is None) draw on the staff seat pool — a
    # client-portal row pins to an existing client, it doesn't create one, so
    # it never touches either seat pool. Tracked locally, same reasoning as
    # commit_clients' seats_left just above in this file.
    staff_seats_left = await remaining_staff_seats(session, user.tenant)

    for index, row in enumerate(rows, start=1):
        email = str(row.email).strip().lower()
        if row.client_id is None and staff_seats_left is not None and staff_seats_left <= 0:
            failed += 1
            errors.append(f"Row {index} ({email}): staff seat limit reached — contact your provider for more seats")
            outcomes.append(
                {"email": email, "full_name": row.full_name, "created": False, "error": "Staff seat limit reached"}
            )
            continue
        try:
            # A savepoint per row: a failure rolls back only that account's
            # profile insert, leaving the successful ones committed.
            async with session.begin_nested():
                result = await accounts.provision(
                    session,
                    tenant=user.tenant,
                    email=email,
                    full_name=row.full_name,
                    role=row.role,
                    client_id=row.client_id,
                    title=row.title,
                    phone=row.phone,
                    send_welcome=send_email,
                    reply_to=user.profile.email,
                )
        except AccountError as exc:
            failed += 1
            errors.append(f"Row {index} ({email}): {exc}")
            outcomes.append({"email": email, "full_name": row.full_name, "created": False, "error": str(exc)})
            continue
        except Exception as exc:  # noqa: BLE001 - surface the row that failed
            failed += 1
            errors.append(f"Row {index} ({email}): {type(exc).__name__}")
            outcomes.append(
                {"email": email, "full_name": row.full_name, "created": False, "error": type(exc).__name__}
            )
            continue

        created += 1
        if row.client_id is None and staff_seats_left is not None:
            staff_seats_left -= 1
        if result.email_sent:
            emailed += 1
        outcomes.append(
            {
                "email": email,
                "full_name": result.profile.full_name,
                "created": True,
                "temp_password": result.temp_password,
                "email_sent": result.email_sent,
            }
        )

    await audit.record(
        session,
        tenant_id=user.tenant_id,
        actor_id=user.profile.id,
        actor_email=user.profile.email,
        action="imported",
        entity="profile",
        summary=f"Imported {created} user account(s)",
        metadata={"created": created, "failed": failed, "emailed": emailed},
    )
    return UserImportResult(
        created=created,
        failed=failed,
        emailed=emailed,
        accounts=outcomes,
        errors=errors[:25],
    )


# -----------------------------------------------------------------------------
# Services — the catalogue a firm assigns to clients
# -----------------------------------------------------------------------------
SERVICE_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "code": ("code", "service code", "short code"),
    "name": ("name", "service", "service name"),
    "description": ("description", "notes", "details"),
    "category": ("category", "group"),
    "frequency": ("frequency", "cadence"),
    "default_price": ("default price", "price", "fee", "fees"),
    "lead_time_days": ("lead time", "lead time days", "lead time (days)"),
    # The due rule is a JSON shape in the DB; a spreadsheet only ever wants
    # the common case — "N months after the fiscal year end" — so that's the
    # one column exposed here. The other rule shape (a fixed calendar date)
    # stays edit-only in the UI, same as it already is.
    "months_after_period_end": (
        "months after period end", "due months", "months after year end", "months",
    ),
    "is_active": ("active", "is active", "status"),
}

FREQUENCY_MAP = {
    "monthly": "monthly",
    "quarterly": "quarterly",
    "semi annual": "semi_annual", "semi-annual": "semi_annual", "biannual": "semi_annual",
    "annual": "annual", "annually": "annual", "yearly": "annual",
    "one time": "one_time", "one-time": "one_time", "once": "one_time", "single": "one_time",
}


def detect_service_mapping(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    for column in columns:
        key = _normalise(column)
        for field, aliases in SERVICE_FIELD_ALIASES.items():
            if field in taken:
                continue
            if key == field or key == field.replace("_", " ") or key in aliases:
                mapping[column] = field
                taken.add(field)
                break
    return mapping


def parse_service_row(raw: dict[str, Any], mapping: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    data: dict[str, Any] = {
        "category": "General",
        "frequency": "annual",
        "default_price": 0.0,
        "lead_time_days": 30,
        "is_active": True,
    }
    errors: list[str] = []
    months_after: int | None = None

    for column, field in mapping.items():
        value = raw.get(column)
        text = "" if value is None else str(value).strip()
        if not text:
            continue

        if field == "code":
            data["code"] = text.upper()
        elif field == "default_price":
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a valid price")
            else:
                data["default_price"] = number
        elif field == "lead_time_days":
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a number for lead time")
            else:
                data["lead_time_days"] = int(number)
        elif field == "frequency":
            resolved = FREQUENCY_MAP.get(text.lower().replace("_", " "))
            if resolved is None:
                errors.append(f"'{text}' is not a frequency we recognise — defaulting to annual")
            else:
                data["frequency"] = resolved
        elif field == "months_after_period_end":
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a number of months")
            else:
                months_after = int(number)
        elif field == "is_active":
            data["is_active"] = text.lower() not in ("no", "false", "0", "inactive", "n")
        else:
            data[field] = text

    data["due_rule"] = {
        "type": "offset_from_period_end",
        "months": months_after if months_after is not None else 6,
        "period_basis": "fiscal",
    }

    if not data.get("code"):
        errors.append("A service code is required")
    if not data.get("name"):
        errors.append("A service name is required")

    return data, errors


@router.post("/services/preview", response_model=ImportPreview)
async def preview_services(
    session: SessionDep, user: TenantUserDep, file: UploadFile = File(...)
) -> ImportPreview:
    columns, records = await _read_table(file)
    mapping = detect_service_mapping(columns)
    if "code" not in mapping.values() or "name" not in mapping.values():
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "No 'Code' and 'Name' columns were recognised. Rename columns to 'Code' and 'Name' and try again.",
        )

    rows: list[ImportPreviewRow] = []
    valid = 0
    seen_codes: set[str] = set()
    for index, record in enumerate(records[:MAX_PREVIEW_ROWS], start=2):
        data, errors = parse_service_row(record, mapping)
        code = str(data.get("code", ""))
        if code:
            if code in seen_codes:
                errors.append("This code is duplicated earlier in the file")
            seen_codes.add(code)
        if not errors:
            valid += 1
        rows.append(ImportPreviewRow(row=index, data=data, errors=errors))

    return ImportPreview(
        columns=columns,
        detected_mapping=mapping,
        rows=rows,
        total_rows=len(records),
        valid_rows=valid,
    )


@router.post("/services/commit", response_model=ImportResult)
async def commit_services(
    payload: ImportCommitRequest, session: SessionDep, user: TenantUserDep
) -> ImportResult:
    created = updated = failed = 0
    errors: list[str] = []
    allowed = {column.name for column in Service.__table__.columns}

    for index, row in enumerate(payload.rows, start=1):
        data = {k: v for k, v in row.items() if k in allowed and k not in ("id", "tenant_id")}
        code = str(data.get("code") or "").strip().upper()
        name = str(data.get("name") or "").strip()
        if not code or not name:
            failed += 1
            errors.append(f"Row {index}: missing code or name")
            continue
        data["code"] = code
        data["name"] = name

        existing = None
        if payload.update_existing:
            existing = await session.scalar(
                select(Service).where(Service.tenant_id == user.tenant_id, Service.code == code)
            )

        try:
            # A savepoint keeps one bad row from discarding the whole import.
            async with session.begin_nested():
                if existing is not None:
                    for key, value in data.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    session.add(Service(tenant_id=user.tenant_id, **data))
                    created += 1
        except Exception as exc:  # noqa: BLE001 - surface the row that failed
            if existing is not None:
                updated -= 1
            else:
                created -= 1
            failed += 1
            errors.append(f"Row {index} ({code}): {type(exc).__name__}")

    await audit.record(
        session,
        tenant_id=user.tenant_id,
        actor_id=user.profile.id,
        actor_email=user.profile.email,
        action="imported",
        entity="service",
        summary=f"Imported {created} new and updated {updated} service(s)",
        metadata={"created": created, "updated": updated, "failed": failed},
    )
    return ImportResult(created=created, updated=updated, failed=failed, errors=errors[:25])


# -----------------------------------------------------------------------------
# Tenants — bulk-provisioning firms (superadmin only)
# -----------------------------------------------------------------------------
TENANT_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("firm name", "name", "company", "company name", "tenant name", "client name"),
    "admin_email": ("admin email", "email", "owner email", "contact email"),
    "admin_name": ("admin name", "owner name", "contact name"),
    "slug": ("slug", "subdomain"),
    "plan": ("plan",),
    "custom_domain": ("custom domain", "domain", "white label domain"),
    "max_clients": ("max clients", "client limit", "client cap"),
    "max_users": ("max users", "user limit", "user cap", "seats"),
    "is_demo": ("demo", "is demo", "sandbox"),
}


def detect_tenant_mapping(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    for column in columns:
        key = _normalise(column)
        for field, aliases in TENANT_FIELD_ALIASES.items():
            if field in taken:
                continue
            if key == field or key == field.replace("_", " ") or key in aliases:
                mapping[column] = field
                taken.add(field)
                break
    return mapping


def parse_tenant_row(raw: dict[str, Any], mapping: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    data: dict[str, Any] = {"plan": "trial", "is_demo": False, "send_email": True}
    errors: list[str] = []

    for column, field in mapping.items():
        value = raw.get(column)
        text = "" if value is None else str(value).strip()
        if not text:
            continue

        if field == "admin_email":
            lowered = text.lower()
            if "@" not in lowered or "." not in lowered.split("@")[-1]:
                errors.append(f"'{text}' is not a valid email")
            else:
                data["admin_email"] = lowered
        elif field in ("max_clients", "max_users"):
            number = _to_float(text)
            if number is None:
                errors.append(f"'{text}' is not a number for {field.replace('_', ' ')}")
            else:
                data[field] = int(number)
        elif field == "is_demo":
            data["is_demo"] = text.lower() in ("yes", "true", "1", "demo", "sandbox")
        elif field == "slug":
            data["slug"] = text.strip().lower()
        else:
            data[field] = text

    if not data.get("name"):
        errors.append("Firm name is required")
    if not data.get("admin_email"):
        errors.append("Admin email is required")

    return data, errors


@router.post("/tenants/preview", response_model=ImportPreview)
async def preview_tenants(
    session: SessionDep, user: SuperadminDep, file: UploadFile = File(...)
) -> ImportPreview:
    columns, records = await _read_table(file)
    mapping = detect_tenant_mapping(columns)
    if "name" not in mapping.values() or "admin_email" not in mapping.values():
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "No 'Firm name' and 'Admin email' columns were recognised. Rename columns and try again.",
        )

    existing_slugs = {row.lower() for row in (await session.scalars(select(Tenant.slug))).all()}

    rows: list[ImportPreviewRow] = []
    valid = 0
    seen_emails: set[str] = set()
    for index, record in enumerate(records[:MAX_PREVIEW_ROWS], start=2):
        data, errors = parse_tenant_row(record, mapping)
        slug = str(data.get("slug", "")).strip().lower()
        if slug and slug in existing_slugs:
            errors.append(f"The slug '{slug}' is already taken")
        email = str(data.get("admin_email", ""))
        if email:
            if email in seen_emails:
                errors.append("This admin email is duplicated earlier in the file")
            seen_emails.add(email)
        if not errors:
            valid += 1
        rows.append(ImportPreviewRow(row=index, data=data, errors=errors))

    return ImportPreview(
        columns=columns,
        detected_mapping=mapping,
        rows=rows,
        total_rows=len(records),
        valid_rows=valid,
    )


@router.post("/tenants/commit", response_model=TenantImportResult)
async def commit_tenants(
    rows: list[TenantAdminCreate], session: SessionDep, user: SuperadminDep, request: Request
) -> TenantImportResult:
    """Provision one firm per row. Each row's failure is independent — a
    taken slug or a bad email on row 3 doesn't stop rows 1, 2 and 4 landing,
    the same reasoning as commit_users below."""
    if len(rows) > MAX_TENANT_COMMIT_ROWS:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Import at most {MAX_TENANT_COMMIT_ROWS} firms at a time — split the file.",
        )

    outcomes: list[TenantImportOutcome] = []
    created = failed = emailed = 0
    errors: list[str] = []
    ip = client_ip(request)

    for index, row in enumerate(rows, start=1):
        try:
            async with session.begin_nested():
                result = await provision_tenant(
                    session,
                    row,
                    actor_id=user.profile.id,
                    actor_email=user.profile.email,
                    ip_address=ip,
                )
        except HTTPException as exc:
            failed += 1
            errors.append(f"Row {index} ({row.name}): {exc.detail}")
            outcomes.append(
                TenantImportOutcome(
                    name=row.name, admin_email=str(row.admin_email), created=False, error=str(exc.detail)
                )
            )
            continue
        except Exception as exc:  # noqa: BLE001 - surface the row that failed
            failed += 1
            errors.append(f"Row {index} ({row.name}): {type(exc).__name__}")
            outcomes.append(
                TenantImportOutcome(
                    name=row.name, admin_email=str(row.admin_email), created=False, error=type(exc).__name__
                )
            )
            continue

        created += 1
        admin = result["admin"]
        if admin["email_sent"]:
            emailed += 1
        outcomes.append(
            TenantImportOutcome(
                name=row.name,
                slug=result["tenant"]["slug"] if isinstance(result["tenant"], dict) else None,
                admin_email=admin["email"],
                created=True,
                temp_password=admin["temp_password"],
                email_sent=admin["email_sent"],
            )
        )

    return TenantImportResult(created=created, failed=failed, emailed=emailed, tenants=outcomes, errors=errors[:25])
